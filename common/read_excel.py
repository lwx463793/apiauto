import openpyxl
import os
import json
from common.constant import EXCEL_DIR
from common.logger import logger



if not os.path.exists(EXCEL_DIR):
    os.mkdir(EXCEL_DIR)


class CasesData:

    def __init__(self, attrs):
        for t in attrs:
            setattr(self, t[0], t[1])


class ReadExcel:

    def __init__(self, file_name, sheet_name):
        """

        :param file_name: 文件地址 -> str
        :param sheet_name: 工作表名字 -> str
        """
        self.file_name = file_name
        self.sheet_name = sheet_name
        # 创建操纵excel的实例
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet_name = self.wb[self.sheet_name]
        logger.info("Workbook: {}, Sheet: {}".format(self.file_name, sheet_name))

    def __del__(self):
        # logger.info("Excel object close")
        self.wb.close()

    def read_excel_dict(self) -> list:
        """
        方法一： 将每条数据以dict的形式储存于list中
        :return: 以列表形式存储的case_data -> list
        """
        # 获取首行数据，得到title并存于list
        title = []
        data = list(self.sheet_name.rows)
        for i in data[0]:
            title.append(i.value)
        # 获取每一行的数据并与title组成dict
        case_data = []
        for r in data[1:]:
            case_list = []
            for va in r:
                try:
                    case_list.append(json.loads(va.value))  # 将json转化成字典
                except (TypeError, SyntaxError, NameError, json.decoder.JSONDecodeError):
                    case_list.append(va.value)
            case_data.append(dict(zip(title, case_list)))  # zip()用于将元素打包成元组，返回迭代器，用list()展示
        print(case_data)
        return case_data



    def read_excel_obj(self, *list_column) -> list:
        """
        自定义选择需要读取的column, 若未传则默认返回全部
        :param list_column: 不定长参数， 需要读取的column(int) -> tuple
        :return: list中存着对象，对象中有实例属性 -> list
        """
        title = []
        if not list_column:  # 未传参时默认为获取全部
            data = list(self.sheet_name.rows)
            for i in data[0]:
                title.append(i.value)
            obj_data = []
            for r in data[1:]:
                case_list = []
                for va in r:
                    try:
                        case_list.append(json.loads(va.value))
                    except (NameError, TypeError, SyntaxError, json.decoder.JSONDecodeError):
                        case_list.append(va.value)
                attr = CasesData(list(zip(title, case_list)))
                print(attr.__dict__)
                obj_data.append(attr)

        else:  # 指定column
            if self.sheet_name.max_column < max(list_column):
                raise ValueError("Column out of range")
            data = list(self.sheet_name.rows)
            for c in list_column:
                title.append(data[0][c - 1].value)
            obj_data = []
            for i in data[1:]:
                case_list = []
                for y in list_column:
                    x = y - 1
                    try:
                        case_list.append(json.loads(i[x].value))
                    except (NameError, TypeError, SyntaxError, json.decoder.JSONDecodeError):
                        case_list.append(i[x].value)
                attr = CasesData(list(zip(title, case_list)))
                obj_data.append(attr)
        return obj_data

    def write_data(self, row, column, msg) -> None:
        """
        写入数据至单元格
        :param row: 行号 -> int
        :param column: 列号 -> int
        :param msg: 数据 -> str
        :return: None
        """
        try:
            logger.info("Excel write to data! row: {}, column: {}, value: {}".format(row, column, msg))
            self.sheet_name.cell(row=row, column=column, value=msg)
            self.wb.save(self.file_name)
        except SyntaxError as e:
            logger.error("Excel Write data function is error: {}".format(e))

